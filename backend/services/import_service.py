"""Здесь опишу логику импорта работников"""
from openpyxl import load_workbook
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from employees.models import Worker


class WorkerImportService:
    """Сервис для импорта работников из Excel."""

    def __init__(self, user):
        self.user = user
        self.created_count = 0
        self.errors = []

    def import_from_excel(self, excel_file):
        """Отвечает за импорт данных из excel."""
        for i, row in self._read_rows(excel_file):
            if not self._validate_required_fields(row, i):
                continue
            if not self._validate_email(row[3], i):
                continue
            if self._is_duplicate(row[3], i):
                continue

            self._create_worker(row)
            self.created_count += 1

        return self._build_report()

    def _read_rows(self, excel_file):
        """Чтение строк из Excel.

        Считывает:
            -строки, начиная со 2-ой без заголовка;
            -столбцы, со 1 по 5.
        """
        wb = load_workbook(excel_file)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            yield i, row[:5]

    def _validate_required_fields(self, row, row_num):
        """Проверка обязательных полей."""
        first_name, middle_name, last_name, email, position = row
        if not (first_name and last_name and email and position):
            self.errors.append(f'Строка {row_num} - отсутствуют обязательные поля')
            return False
        return True

    def _validate_email(self, email, row_num):
        """Проверка корректности email."""
        try:
            validate_email(email)
        except ValidationError:
            self.errors.append(f'Строка {row_num} - некорректный email ({email})')
            return False
        return True

    def _is_duplicate(self, email, row_num):
        """Проверка, что email не дублируется в БД."""
        if Worker.objects.filter(email=email).exists():
            self.errors.append(f'Строка {row_num} - содержит дубликат email ({email})')
            return True
        return False

    def _create_worker(self, row):
        """Создание работника в БД."""
        first_name, middle_name, last_name, email, position = row
        Worker.objects.create(
            first_name=first_name,
            middle_name=middle_name,
            last_name=last_name,
            email=email,
            position=position,
            created_by=self.user,
        )

    def _build_report(self):
        """Формирование отчёта по результатам импорта."""
        return {
            'created': self.created_count,
            'errors_count': len(self.errors),
            'errors': self.errors,
        }
